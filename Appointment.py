def appointment_confirmation(uid):
    import openpyxl as xl

    wb = xl.load_workbook('Appointments.xlsx')
    sheet = wb['Sheet1']
    comment = ''
    column = 1
    while column <= sheet.max_column:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            cell_val = cell.value
            if cell_val == 'NA':
                cell.value = uid
                doc_cell = sheet.cell(row, 1)
                doc = doc_cell.value
                time_cell = sheet.cell(1, column)
                time = time_cell.value
                comment = f'Your appointment is confirmed with {doc} at {time}'
                break
        column += 1
        if cell.value == uid:
            break
    wb.save('Appointments.xlsx')
    return comment
