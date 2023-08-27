def body_mass_index(height, mass):  # height should be in meters and mass in Kg
    height = float(height); mass = float(mass)

    bmi = mass / (height ** 2)
    if (bmi >= 18.5) and (bmi <= 24.9):
        comment = f'Weight is normal. BMI is {bmi}'
        return comment
    elif (bmi >= 25) and (bmi <= 29.9):
        comment = f'Patient is Overweight. BMI is {bmi}'
        return comment
    elif bmi < 18.5:
        comment = f'Patient is Underweight. BMI is {bmi}'
        return comment
    elif bmi >= 30:
        comment = f'The patient is Obese. BMI is {bmi}'
        return comment


def blood_pressure(systole, diastole):
    systole = float(systole); diastole=float(diastole)

    if (systole <= 120) and (diastole <= 80):
        comment = 'Blood Pressure is normal'
        return comment
    elif ((systole > 120) and (systole <= 129)) and (diastole < 80):
        comment = 'Elevated blood pressure'
        return comment
    elif ((systole >= 130) and (systole <= 139)) or ((diastole >= 80) and (diastole <= 89)):
        comment = 'Stage 1 high blood pressure (hypertension)'
        return comment
    elif (systole >= 140) or (diastole >= 90):
        comment = 'Stage 2 high blood pressure (hypertension)'
        return comment


def recorded_symptoms(symptom):
    import openpyxl as xl
    workbook = xl.load_workbook('symptoms.xlsx')
    sheet = workbook['Sheet1']
    row = sheet.max_row
    for row_num in range(2, row + 1):
        cell = sheet.cell(row_num, 1)
        sym = cell.value
        if sym == symptom:
            cell = sheet.cell(row_num, 2)
            med = cell.value
            comment = f'Medicines for {symptom} are {med}'
            return comment


def doctor_prescription():
    import speech_recognition

    voice = speech_recognition.Recognizer()
    b = ''
    with speech_recognition.Microphone() as source:
        print('You can record the audio now')
        try:

            voice.adjust_for_ambient_noise(source)
            try:
                a = voice.record(source, 5)
                b = voice.recognize_google(a)

            except speech_recognition.exceptions.RequestError:
                print("Can't understand you.")

            print('you said', b)
        except speech_recognition.exceptions.UnknownValueError or speech_recognition.exceptions.WaitTimeoutError:
            print("Mic Issues")
        return b


def new_symptom(symptom, medicine):
    import openpyxl as xl

    wb = xl.load_workbook('symptoms.xlsx')
    sheet = wb['Sheet1']
    max_row = sheet.max_row
    symptom_cell = sheet.cell(max_row + 2, 1)
    symptom_cell.value = symptom
    medicine_cell = sheet.cell(max_row + 2, 2)
    medicine_cell.value = medicine

    wb.save('symptoms.xlsx')

